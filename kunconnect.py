import os
import subprocess
import glob
from openpyxl import load_workbook

#-----------------------------------------------------------------------------#
#
#                      執行YoloV5偵測樓層跨數、樓層數
#                             
#-----------------------------------------------------------------------------#

# 要執行的命令及參數
command = ["python", "./detect.py","--weights", "runs/train/exp2/weights/best.pt","--source","data/images/8.png", "--save-txt"]

# 使用 subprocess 模組執行命令
process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)

# 等待命令執行完成
stdout, stderr = process.communicate()

# 檢查執行結果
if process.returncode == 0:
    print("執行成功！")
    print("輸出：", stdout.decode())
else:
    print("執行失敗！")
    print("錯誤信息：", stderr.decode())

# 定義結果保存的根目錄
results_dir = "runs/detect/"

# 列出目錄中的所有文件和文件夾
contents = os.listdir(results_dir)

# 找到最新生成的目錄
latest_exp_dir = max(contents, key=lambda x: os.path.getctime(os.path.join(results_dir, x)))

#-----------------------------------------------------------------------------#
#
#                      將輸出結果進行後處理
#                             
#-----------------------------------------------------------------------------#

# 使用 os.path.join 組合路徑
folder_path = os.path.join("./runs", "detect", latest_exp_dir, "labels")
search_path = os.path.join(folder_path, "*.txt")

# 列出所有 .txt 文件
all_txt_files = glob.glob(search_path)

for txt_file in all_txt_files:
    with open(txt_file, "r") as file:
        lines = file.readlines()

    print("訓練結果儲存於",txt_file)

    # 儲存 bounding box 的 x_center 座標
    x_centers = []
    y_centers = []

    # 解析每一行標註
    for line in lines:
        class_label, x_center, y_center, width, height = map(float, line.split())

        x_centers.append(x_center)
        y_centers.append(y_center)


    # 將 y_center 座標排序
    y_centers.sort()
    x_centers.sort()

    # 初始化計數器
    x_group_count = 0
    y_group_count = 0

    x_group_tolerance = 0.2  # x分組容忍度
    y_group_tolerance = 0.02  # y分組容忍度

    x_current_group = [x_centers[0]]  # 目前的分組
    y_current_group = [y_centers[0]]  # 目前的分組

    # 檢查 x_center 座標是否接近彼此的 bounding box
    for i in range(1, len(x_centers)):
        if abs(x_centers[i] - x_current_group[-1]) <= x_group_tolerance:
            x_current_group.append(x_centers[i])  # 將 x_center 加入目前的分組
        else:
            x_group_count += 1

            # print(f"Group {x_group_count}: {x_current_group}")  # 印出目前的分組

            x_current_group = [x_centers[i]]  # 開始新的分組

    # 印出最後一組
    if x_current_group:
        x_group_count += 1
        
        # print(f"Group {x_group_count}: {x_current_group}")

    # print("共有", x_group_count, "組 bounding box 的 x_center 座標接近。")

    # 檢查 y_center 座標是否接近彼此的 bounding box
    for i in range(1, len(y_centers)):
        if abs(y_centers[i] - y_current_group[-1]) <= y_group_tolerance:
            y_current_group.append(y_centers[i])  # 將 x_center 加入目前的分組
        else:
            y_group_count += 1
            
            # print(f"Group {y_group_count}: {y_current_group}")  # 印出目前的分組
            
            y_current_group = [y_centers[i]]  # 開始新的分組

    # 印出最後一組
    if y_current_group:
        y_group_count += 1
        
        # print(f"Group {y_group_count}: {y_current_group}")
        
    # print("共有", y_group_count, "組 bounding box 的 y_center 座標接近。")

    print("此建築為",x_group_count,"跨",y_group_count,"層")

#-----------------------------------------------------------------------------#
#
#                      修改坤哥建築資訊 Excel 輸入文件
#                             
#-----------------------------------------------------------------------------#

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# 读取 Excel 文件
file_path = r"C:\\Users\\user\\Desktop\\Data_Base_kun\\輸入模型資料.xlsx"

# 使用 pandas 读取目标工作表的数据
df = pd.read_excel(file_path, sheet_name='建築資訊')

df.iloc[0, 1] = y_group_count  # 修改楼层数 (B2, 注意 iloc 是从 0 开始计数)
df.iloc[1, 1] = x_group_count  # 修改跨数 (B3)

# 保存修改后的 DataFrame 到临时 Excel 文件
temp_file_path = r"C:\\Users\\user\\Desktop\\Data_Base_kun\\temp.xlsx"
df.to_excel(temp_file_path, sheet_name='建築資訊', index=False)

# 使用 openpyxl 载入原始工作簿和临时工作簿
workbook = load_workbook(file_path)
temp_workbook = load_workbook(temp_file_path)

# 获取原始工作簿和临时工作簿中的工作表
sheet = workbook['建築資訊']
temp_sheet = temp_workbook['建築資訊']

# 将临时工作表中的数据写入原始工作表，同时保持格式
for row in temp_sheet.iter_rows(min_row=1, max_row=temp_sheet.max_row, min_col=1, max_col=temp_sheet.max_column):
    for cell in row:
        # 跳过合并单元格
        if isinstance(sheet[cell.coordinate], MergedCell):
            continue
        sheet[cell.coordinate].value = cell.value

# 删除临时文件
import os
os.remove(temp_file_path)

# 保存修改后的工作簿
workbook.save(file_path)

print("Excel 文件更新成功！")


#-----------------------------------------------------------------------------#
#
#                              執行坤哥程式碼
#                             
#-----------------------------------------------------------------------------#

# 要執行的命令及參數
command = ["python", "C:/Users/user/Desktop/Data_Base_kun/Python_MainProgram.py"]

# 指定工作目录为其他文件夹的路径
working_directory = "C:/Users/user/Desktop/Data_Base_kun/"

# 使用 subprocess 模块执行命令，并指定工作目录
process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, cwd=working_directory)

# 等待命令執行完成
stdout, stderr = process.communicate()

# 檢查執行結果
if process.returncode == 0:
    print("執行成功！")
    print("輸出：", stdout.decode())
else:
    print("執行失敗！")
    try:
        print("错误信息：", stderr.decode(errors='ignore'))  # 忽略无法解码的字节
    except Exception as e:
        print("无法解码错误信息：", e)